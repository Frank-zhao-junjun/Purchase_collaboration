"""供应商门户协同API - Phase 1: 订单协同闭环"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, update, func
from sqlalchemy.orm import selectinload
from typing import List, Optional
from datetime import datetime
import uuid, json

from app.core.database import get_db
from app.models.supply_chain import (
    Supplier, PurchaseOrder, PurchaseOrderItem, Material,
    ShipmentNote, ShipmentNoteItem, OrderStatus, ShipmentStatus,
    PurchaseForecast, ForecastStatus
)
from app.schemas.supply_chain import (
    POConfirmRequest, ShipmentNoteCreate, ShipmentNoteUpdate,
    ShipmentNoteResponse, ShipmentNoteItemCreate,
    PurchaseOrderExtendedResponse, SupplierExtendedResponse,
    SupplierExtendedUpdate, SupplierTodoStats
)

router = APIRouter(prefix="/supplier-portal", tags=["供应商门户-协同"])


# ==================== 1. 供应商工作台 ====================

@router.get("/todo-stats", response_model=SupplierTodoStats)
async def get_supplier_todo_stats(supplier_id: int, db: AsyncSession = Depends(get_db)):
    """获取供应商待办统计 - 门户首页核心数据"""
    # 待确认订单 (pending 状态的 PO)
    po_result = await db.execute(
        select(func.count()).select_from(PurchaseOrder)
        .where(PurchaseOrder.supplier_id == supplier_id)
        .where(PurchaseOrder.status == OrderStatus.PENDING)
        .where(PurchaseOrder.supplier_confirmed_at.is_(None))
    )
    pending_confirm = po_result.scalar() or 0

    # 待发货 (已确认但未创建ASN或ASN还在draft状态)
    shipment_result = await db.execute(
        select(func.count()).select_from(ShipmentNote)
        .where(ShipmentNote.supplier_id == supplier_id)
        .where(ShipmentNote.status.in_([ShipmentStatus.DRAFT, ShipmentStatus.SUBMITTED]))
    )
    pending_shipment = shipment_result.scalar() or 0

    # 运输中
    transit_result = await db.execute(
        select(func.count()).select_from(ShipmentNote)
        .where(ShipmentNote.supplier_id == supplier_id)
        .where(ShipmentNote.status == ShipmentStatus.IN_TRANSIT)
    )
    in_transit = transit_result.scalar() or 0

    return SupplierTodoStats(
        pending_confirm_orders=pending_confirm,
        pending_shipment_notes=pending_shipment,
        in_transit_count=in_transit,
        pending_settlement=0  # Phase 2 实现
    )


@router.get("/orders", response_model=List[PurchaseOrderExtendedResponse])
async def list_supplier_orders(
    supplier_id: int,
    status: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
):
    """供应商查看自己的采购订单列表"""
    query = (
        select(PurchaseOrder)
        .options(selectinload(PurchaseOrder.items))
        .options(selectinload(PurchaseOrder.supplier))
        .where(PurchaseOrder.supplier_id == supplier_id)
        .order_by(PurchaseOrder.created_at.desc())
    )
    if status:
        query = query.where(PurchaseOrder.status == status)
    result = await db.execute(query)
    orders = result.scalars().all()

    # 附加 supplier_name
    response = []
    for o in orders:
        d = o.__dict__.copy()
        d['supplier_name'] = o.supplier.name if o.supplier else None
        response.append(d)
    return response


# ==================== 2. PO 确认/拒绝协同 ====================

@router.post("/orders/{order_id}/confirm")
async def confirm_or_reject_order(
    order_id: int,
    body: POConfirmRequest,
    supplier_id: int,
    db: AsyncSession = Depends(get_db)
):
    """供应商确认或拒绝采购订单"""
    result = await db.execute(
        select(PurchaseOrder).options(selectinload(PurchaseOrder.items))
        .where(PurchaseOrder.id == order_id)
        .where(PurchaseOrder.supplier_id == supplier_id)
    )
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在或不属于该供应商")

    if order.supplier_confirmed_at is not None:
        raise HTTPException(status_code=400, detail="该订单已处理，无法重复操作")

    now = datetime.utcnow()
    if body.action == "confirm":
        order.status = OrderStatus.CONFIRMED
        order.supplier_confirmed_at = now
    elif body.action == "reject":
        order.status = OrderStatus.CANCELLED
        order.supplier_confirmed_at = now
        order.supplier_rejection_reason = body.rejection_reason or "供应商拒绝"
    else:
        raise HTTPException(status_code=400, detail="action 必须为 confirm 或 reject")

    await db.flush()
    await db.refresh(order)
    return {
        "success": True,
        "message": f"订单已{'确认' if body.action == 'confirm' else '拒绝'}",
        "order_id": order.id,
        "status": order.status.value
    }


# ==================== 3. ASN 送货单协同 ====================

def generate_shipment_no():
    return f"ASN{datetime.now().strftime('%Y%m%d')}{uuid.uuid4().hex[:6].upper()}"


@router.get("/shipments", response_model=List[ShipmentNoteResponse])
async def list_supplier_shipments(
    supplier_id: int,
    status: Optional[str] = None,
    purchase_order_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db)
):
    """供应商查看自己的送货单列表"""
    query = (
        select(ShipmentNote)
        .options(selectinload(ShipmentNote.items))
        .options(selectinload(ShipmentNote.purchase_order))
        .options(selectinload(ShipmentNote.supplier))
        .where(ShipmentNote.supplier_id == supplier_id)
        .order_by(ShipmentNote.created_at.desc())
    )
    if status:
        query = query.where(ShipmentNote.status == status)
    if purchase_order_id:
        query = query.where(ShipmentNote.purchase_order_id == purchase_order_id)

    result = await db.execute(query)
    shipments = result.scalars().all()

    response = []
    for s in shipments:
        d = s.__dict__.copy()
        d['purchase_order_no'] = s.purchase_order.order_no if s.purchase_order else None
        d['supplier_name'] = s.supplier.name if s.supplier else None
        response.append(d)
    return response


@router.get("/shipments/{shipment_id}", response_model=ShipmentNoteResponse)
async def get_shipment_detail(
    shipment_id: int,
    supplier_id: int,
    db: AsyncSession = Depends(get_db)
):
    """查看送货单详情"""
    result = await db.execute(
        select(ShipmentNote)
        .options(selectinload(ShipmentNote.items))
        .options(selectinload(ShipmentNote.purchase_order))
        .options(selectinload(ShipmentNote.supplier))
        .where(ShipmentNote.id == shipment_id)
        .where(ShipmentNote.supplier_id == supplier_id)
    )
    sn = result.scalar_one_or_none()
    if not sn:
        raise HTTPException(status_code=404, detail="送货单不存在")

    d = sn.__dict__.copy()
    d['purchase_order_no'] = sn.purchase_order.order_no if sn.purchase_order else None
    d['supplier_name'] = sn.supplier.name if sn.supplier else None
    return d


@router.post("/shipments", response_model=ShipmentNoteResponse, status_code=201)
async def create_shipment_note(
    data: ShipmentNoteCreate,
    supplier_id: int,
    db: AsyncSession = Depends(get_db)
):
    """供应商创建送货单(ASN)"""
    # 验证 PO 存在且属于该供应商
    po_result = await db.execute(
        select(PurchaseOrder).where(PurchaseOrder.id == data.purchase_order_id)
        .where(PurchaseOrder.supplier_id == supplier_id)
    )
    po = po_result.scalar_one_or_none()
    if not po:
        raise HTTPException(status_code=404, detail="采购订单不存在或不属于该供应商")

    total_qty = sum(item.quantity for item in data.items)
    sn = ShipmentNote(
        shipment_no=generate_shipment_no(),
        purchase_order_id=data.purchase_order_id,
        supplier_id=supplier_id,
        status=ShipmentStatus.DRAFT,
        carrier_name=data.carrier_name,
        tracking_no=data.tracking_no,
        vehicle_no=data.vehicle_no,
        driver_name=data.driver_name,
        driver_phone=data.driver_phone,
        expected_arrival=data.expected_arrival,
        shipping_address=data.shipping_address,
        receiving_warehouse=data.receiving_warehouse,
        notes=data.notes,
        total_quantity=total_qty
    )
    db.add(sn)
    await db.flush()

    for item in data.items:
        # 获取物料名称
        mat_result = await db.execute(select(Material).where(Material.id == item.material_id))
        material = mat_result.scalar_one_or_none()
        mat_name = material.name if material else item.material_name or "未知物料"

        db.add(ShipmentNoteItem(
            shipment_note_id=sn.id,
            material_id=item.material_id,
            material_name=mat_name,
            quantity=item.quantity,
            unit=item.unit,
            batch_no=item.batch_no,
            production_date=item.production_date,
            origin_location=item.origin_location,
            quality_grade=item.quality_grade,
            package_count=item.package_count or 1,
            notes=item.notes
        ))
    await db.flush()

    # 返回完整数据
    result = await db.execute(
        select(ShipmentNote)
        .options(selectinload(ShipmentNote.items))
        .options(selectinload(ShipmentNote.purchase_order))
        .options(selectinload(ShipmentNote.supplier))
        .where(ShipmentNote.id == sn.id)
    )
    created = result.scalar_one()
    d = created.__dict__.copy()
    d['purchase_order_no'] = created.purchase_order.order_no if created.purchase_order else None
    d['supplier_name'] = created.supplier.name if created.supplier else None
    return d


@router.put("/shipments/{shipment_id}/submit")
async def submit_shipment_note(
    shipment_id: int,
    supplier_id: int,
    db: AsyncSession = Depends(get_db)
):
    """提交送货单（从草稿→已提交）"""
    result = await db.execute(
        select(ShipmentNote).where(ShipmentNote.id == shipment_id)
        .where(ShipmentNote.supplier_id == supplier_id)
    )
    sn = result.scalar_one_or_none()
    if not sn:
        raise HTTPException(status_code=404, detail="送货单不存在")
    if sn.status != ShipmentStatus.DRAFT:
        raise HTTPException(status_code=400, detail="只有草稿状态的送货单可以提交")

    sn.status = ShipmentStatus.SUBMITTED
    # 同时更新 PO 状态为已发货
    await db.execute(
        update(PurchaseOrder)
        .where(PurchaseOrder.id == sn.purchase_order_id)
        .values(status=OrderStatus.SHIPPED)
    )
    await db.flush()
    return {"success": True, "message": "送货单已提交", "shipment_no": sn.shipment_no}


# ==================== 4. 收货/验收结果查询（供应商视角）====================

@router.get("/receipts")
async def list_supplier_receipts(
    supplier_id: int,
    status: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
):
    """供应商查看收货验收结果 - 基于送货单状态"""
    query = (
        select(ShipmentNote)
        .options(selectinload(ShipmentNote.items))
        .options(selectinload(ShipmentNote.purchase_order))
        .where(ShipmentNote.supplier_id == supplier_id)
        .where(ShipmentNote.status.in_([ShipmentStatus.ARRIVED, ShipmentStatus.RECEIVED]))
        .order_by(ShipmentNote.updated_at.desc())
    )
    result = await db.execute(query)
    receipts = result.scalars().all()

    response = []
    for r in receipts:
        d = r.__dict__.copy()
        d['purchase_order_no'] = r.purchase_order.order_no if r.purchase_order else None
        items_data = []
        for item in r.items:
            items_data.append({
                "id": item.id,
                "material_name": item.material_name,
                "quantity": item.quantity,
                "batch_no": item.batch_no,
                "quality_grade": item.quality_grade
            })
        d['items'] = items_data
        response.append(d)
    return response


# ==================== 5. 供应商主数据自维护 ====================

@router.get("/profile/{supplier_id}", response_model=SupplierExtendedResponse)
async def get_supplier_profile(supplier_id: int, db: AsyncSession = Depends(get_db)):
    """供应商查看自己的完整资料"""
    result = await db.execute(select(Supplier).where(Supplier.id == supplier_id))
    supplier = result.scalar_one_or_none()
    if not supplier:
        raise HTTPException(status_code=404, detail="供应商不存在")
    return supplier


@router.put("/profile/{supplier_id}", response_model=SupplierExtendedResponse)
async def update_supplier_profile(
    supplier_id: int,
    data: SupplierExtendedUpdate,
    db: AsyncSession = Depends(get_db)
):
    """供应商自助编辑企业信息"""
    result = await db.execute(select(Supplier).where(Supplier.id == supplier_id))
    supplier = result.scalar_one_or_none()
    if not supplier:
        raise HTTPException(status_code=404, detail="供应商不存在")

    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(supplier, key, value)
    await db.flush()
    await db.refresh(supplier)
    return supplier


@router.post("/profile/{supplier_id}/qualifications")
async def add_qualification(
    supplier_id: int,
    qualification: dict,
    db: AsyncSession = Depends(get_db)
):
    """添加资质证书"""
    result = await db.execute(select(Supplier).where(Supplier.id == supplier_id))
    supplier = result.scalar_one_or_none()
    if not supplier:
        raise HTTPException(status_code=404, detail="供应商不存在")

    try:
        quals = json.loads(supplier.qualifications or "[]")
    except Exception:
        quals = []

    new_qual = {
        "id": len(quals) + 1,
        "name": qualification.get("name", ""),
        "type": qualification.get("type", ""),
        "cert_no": qualification.get("cert_no", ""),
        "issue_date": str(qualification.get("issue_date", "")),
        "expire_date": str(qualification.get("expire_date", "")),
        "status": qualification.get("status", "valid"),
        "added_at": datetime.utcnow().isoformat()
    }
    quals.append(new_qual)
    supplier.qualifications = json.dumps(quals, ensure_ascii=False)
    await db.flush()
    return {"success": True, "qualification": new_qual}


@router.get("/profile/{supplier_id}/qualifications")
async def list_qualifications(supplier_id: int, db: AsyncSession = Depends(get_db)):
    """列出所有资质证书"""
    result = await db.execute(select(Supplier).where(Supplier.id == supplier_id))
    supplier = result.scalar_one_or_none()
    if not supplier:
        raise HTTPException(status_code=404, detail="供应商不存在")
    try:
        return json.loads(supplier.qualifications or "[]")
    except Exception:
        return []

# ==================== 6. 采购预测 (US-301) ====================

@router.get("/forecasts")
async def list_forecasts(
    supplier_id: Optional[int] = None,
    status: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
):
    """获取采购预测列表（采购端/供应商端通用）"""
    query = (
        select(PurchaseForecast)
        .options(selectinload(PurchaseForecast.supplier))
        .order_by(PurchaseForecast.created_at.desc())
    )
    if supplier_id:
        query = query.where(PurchaseForecast.supplier_id == supplier_id)
    if status:
        query = query.where(PurchaseForecast.status == status)

    result = await db.execute(query)
    forecasts = result.scalars().all()

    response = []
    for f in forecasts:
        d = {c.name: getattr(f, c.name) for c in f.__table__.columns}
        d['supplier_name'] = f.supplier.name if f.supplier else None
        response.append(d)
    return response


@router.post("/forecasts/import-excel")
async def import_forecast_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    """通过 Excel 导入采购预测数据"""
    import openpyxl
    from io import BytesIO

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 文件")

    contents = await file.read()
    wb = openpyxl.load_workbook(BytesIO(contents))
    ws = wb.active

    if ws.max_row < 2:
        raise HTTPException(status_code=400, detail="Excel 文件为空，请至少包含表头和数据行")

    headers = [str(c.value or "").strip() for c in ws[1]]
    required = ["物料名称", "预测数量"]
    for r in required:
        if not any(r in h for h in headers):
            raise HTTPException(status_code=400, detail=f"缺少必要列: {r}")

    def col(name):
        for i, h in enumerate(headers):
            if name in h:
                return i
        return -1

    idx_material = col("物料名称")
    idx_material_id = col("物料ID")
    idx_month = col("预测月份")
    idx_qty = col("预测数量")
    idx_notes = col("备注")

    if idx_material < 0 or idx_qty < 0:
        raise HTTPException(status_code=400, detail="Excel 必须包含「物料名称」和「预测数量」列")

    from datetime import datetime

    rows_data = []
    errors = []

    for row_idx in range(2, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        material_name = str(row[idx_material] or "").strip()
        qty_val = row[idx_qty]

        if not material_name:
            continue

        try:
            qty = float(qty_val) if qty_val is not None else 0
        except (ValueError, TypeError):
            errors.append(f"第{row_idx}行: 预测数量「{qty_val}」不是有效数字")
            continue

        material_id = None
        if idx_material_id >= 0 and row[idx_material_id] is not None:
            try:
                material_id = int(float(row[idx_material_id]))
            except (ValueError, TypeError):
                pass

        month_str = str(row[idx_month] or "").strip() if idx_month >= 0 else ""
        notes = str(row[idx_notes] or "").strip() if idx_notes >= 0 else ""

        rows_data.append({
            "material_id": material_id,
            "material_name": material_name,
            "quantity": qty,
            "expected_month": month_str,
            "notes": notes,
        })

    if not rows_data:
        raise HTTPException(status_code=400, detail="Excel 中无有效数据行")

    supplier_result = await db.execute(
        select(Supplier).where(Supplier.status == "active")
    )
    suppliers = supplier_result.scalars().all()

    if not suppliers:
        raise HTTPException(status_code=400, detail="没有活跃的供应商，无法创建预测")

    now = datetime.utcnow()
    created_count = 0

    for supplier in suppliers:
        forecast = PurchaseForecast(
            supplier_id=supplier.id,
            forecast_period=f"{now.year}年{now.month}月",
            period_start=datetime(now.year, now.month, 1),
            period_end=datetime(now.year + (now.month // 12), (now.month % 12) + 1, 1) if now.month < 12 else datetime(now.year + 1, 1, 1),
            items_data=json.dumps(rows_data, ensure_ascii=False),
            status="published",
            created_at=now,
            published_at=now,
        )
        db.add(forecast)
        created_count += 1

    await db.commit()

    return {
        "success": True,
        "message": f"成功从 Excel 导入 {len(rows_data)} 条预测明细，已分发给 {created_count} 个供应商",
        "items_count": len(rows_data),
        "suppliers_count": created_count,
        "errors": errors if errors else None,
    }


@router.get("/forecasts/{forecast_id}")
async def get_forecast_detail(
    forecast_id: int,
    db: AsyncSession = Depends(get_db)
):
    """获取采购预测详情"""
    result = await db.execute(
        select(PurchaseForecast)
        .options(selectinload(PurchaseForecast.supplier))
        .where(PurchaseForecast.id == forecast_id)
    )
    f = result.scalar_one_or_none()
    if not f:
        raise HTTPException(status_code=404, detail="预测记录不存在")

    d = {c.name: getattr(f, c.name) for c in f.__table__.columns}
    d['supplier_name'] = f.supplier.name if f.supplier else None
    return d
